#!/usr/bin/env python3
"""
Build the O8FFL static-site data layer.

Reads the league master workbook (.xlsx) and the by-laws (.docx) and writes JSON
(plus extracted rulebook images) into ./data and ./assets/rulebook.

Design notes
------------
* Nothing is hardcoded to a particular season. Draft sheets are discovered by the
  pattern "<year> Draft" and roster sheets by a bare "<year>" sheet name, so adding
  a "2026 Draft" tab (and a "2025" roster tab) next summer just works.
* Owner names in the workbook are inconsistent (first names on draft boards, typos,
  franchises that changed hands). Everything is normalised onto a franchise (team
  number) + a per-season owner display name. Unrecognised names are reported at the
  end of the run instead of failing silently.

Usage:  python3 scripts/build_data.py [--workbook PATH] [--bylaws PATH] [--out DIR]
"""

from __future__ import annotations

import argparse
import base64
import difflib
import glob
import hashlib
import json
import math
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  python3 -m pip install --user openpyxl python-docx")

try:
    import docx as python_docx
except ImportError:  # pragma: no cover
    python_docx = None


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# --------------------------------------------------------------------------------------
# Owner / franchise registry
# --------------------------------------------------------------------------------------
# Aliases the workbook uses for a franchise's owners. Keys are matched after
# normalisation (lowercase, punctuation and whitespace stripped), so "C.J.", "CJ" and
# "c j" all collapse to the same key. Add a line here if a new owner or nickname shows up.
#
# The site is public, so nobody's surname is published: every owner is shown by first name
# only. Current owners get a bare first name; a former owner who shares a first name with
# a current one keeps a last initial (Jon F., Ryan A.) so old seasons stay unambiguous.
OWNER_ALIASES = {
    # team number: (full name — never published, display name, [aliases])
    1: ("C.J. Bradford", "C.J.", ["cj", "cjbradford", "charlesbradford", "bradford", "cj bradford"]),
    2: ("Chris Dooley", "Chris", ["chris", "chrisdooely", "chrisdooley", "cdooley"]),
    3: ("Connor Shea", "Connor", ["connor", "connorshea", "shea"]),
    4: ("Brian Dooley", "Brian", ["brian", "brain", "briandooley", "bdooley"]),
    5: ("Michael Caruso", "Mike", ["mike", "caruso", "mikecaruso", "michaelcaruso"]),
    6: ("Matt Rose", "Matt", ["matt", "matthewrose", "mattrose", "rose"]),
    7: ("Ryan Perkins", "Ryan", ["ryan", "ryanperkins", "perkins"]),
    8: ("Shawn Cronin", "Shawn", ["shawn", "shawncronin", "cronin", "sean"]),
    9: ("Taylor Whitcomb", "Taylor", ["taylor", "taylorwhitcomb", "whitcomb"]),
    # "Jon" defaults to Jon Sanders; seasons Jon Foster played resolve by season below.
    10: ("Jon Sanders", "Jon", ["sanders", "jonsanders", "johnsanders", "jsanders", "jon", "john"]),
}

# Former owners of a franchise. These people no longer play; they keep their own display
# name in historical seasons but map to the same franchise/team number.
FORMER_OWNERS = {
    4: [("Dan Gordon", "Dan")],
    5: [("Ryan Aberdale", "Ryan A."), ("Jon Foster", "Jon F.")],
    10: [("Tyler Moules", "Tyler")],
}

# Surnames safe to strip on sight in free text. Deliberately excludes ones that collide
# with ordinary words or trophy names — "Rose", "Dooley", "Moules" (the trophy), "Foster",
# "Gordon" — those are only replaced as part of a full name.
# Only applied to fantasy team names, where the surrounding words make a surname
# unambiguous ("TEAM FOSTER"). Too risky for prose.
TEAM_NAME_SURNAMES = {"Foster": "Jon F.", "Gordon": "Dan", "Moules": "Tyler"}

# Surnames as they get misspelled in the meeting minutes and draft boards.
TYPO_SURNAMES = {"Whitcomb": ["Witcomb", "Whitcombe"], "Caruso": ["Caurso"], "Dooley": ["Dooely"],
                 "Bradford": ["Bradfrod"], "Perkins": ["Perkings"]}

LONE_SURNAMES = {
    "Bradford": "C.J.",
    "Shea": "Connor",
    "Caruso": "Mike",
    "Perkins": "Ryan",
    "Cronin": "Shawn",
    "Whitcomb": "Taylor",
    "Sanders": "Jon",
    "Aberdale": "Ryan A.",
}

# Ambiguous short aliases that can only be resolved with the season in hand
# (e.g. "Jon" means Jon Foster in 2014-2016 and Jon Sanders afterwards). These are
# resolved against the set of owners known to have played that season.
AMBIGUOUS_FIRST_NAMES = {"jon", "john", "ryan", "chris", "dan", "tyler", "mike", "brian"}

POSITIONS = ("QB", "RB", "WR", "TE", "K", "D/ST", "DST", "PK", "P", "FLEX", "Bench", "IR")

NAME_SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v"}


def norm_key(value) -> str:
    """Aggressive normalisation used for name matching only (never for display)."""
    if value is None:
        return ""
    s = unicodedata.normalize("NFKD", str(value))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s.lower())


class OwnerRegistry:
    """Maps every spelling of an owner in the workbook onto a franchise."""

    def __init__(self):
        self.by_key: dict[str, dict] = {}
        self.teams: dict[int, dict] = {}
        self.unresolved: Counter = Counter()
        self.season_owners: dict[int, set[str]] = defaultdict(set)  # year -> canonical names

        self.full_to_display: dict[str, str] = {}

        for team, (full, display, aliases) in OWNER_ALIASES.items():
            self.teams[team] = {
                "team": team,
                "owner": display,
                "former_owners": [short for _, short in FORMER_OWNERS.get(team, [])],
            }
            self.full_to_display[full] = display
            self._register(full, team, display)
            self._register(display, team, display)
            for alias in aliases:
                self._register(alias, team, display)
            for former_full, former_display in FORMER_OWNERS.get(team, []):
                self.full_to_display[former_full] = former_display
                self._register(former_full, team, former_display)
                self._register(former_display, team, former_display)
                # last name alone, e.g. "Aberdale", "Moules", "Gordon", "Foster"
                self._register(former_full.split()[-1], team, former_display)

    def _register(self, alias, team, display):
        key = norm_key(alias)
        if not key:
            return
        # First registration wins so canonical names beat generic first names.
        self.by_key.setdefault(key, {"team": team, "display": display})

    def note_season(self, year: int, canonical_name: str):
        self.season_owners[year].add(canonical_name)

    def resolve(self, raw, year=None, record=True):
        """Return {'team': int, 'display': owner name} or None."""
        key = norm_key(raw)
        if not key:
            return None

        # A first name that belonged to more than one person: prefer the owner who
        # actually played that season.
        if year and key in AMBIGUOUS_FIRST_NAMES:
            candidates = []
            for name in self.season_owners.get(year, ()):  # canonical season names
                first = norm_key(name.split()[0])
                if first == key or first.startswith(key) or key.startswith(first):
                    candidates.append(name)
            if len(candidates) == 1:
                hit = self.by_key.get(norm_key(candidates[0]))
                if hit:
                    return dict(hit)

        hit = self.by_key.get(key)
        if hit:
            return dict(hit)

        # Fall back to a unique prefix/substring match against known spellings.
        matches = {v["team"]: v for k, v in self.by_key.items() if k.startswith(key) or key.startswith(k)}
        if len(matches) == 1:
            return dict(next(iter(matches.values())))

        if record:
            self.unresolved[str(raw).strip()] += 1
        return None

    def owner_for_team(self, team):
        entry = self.teams.get(team)
        return entry["owner"] if entry else None

    def shorten_team_name(self, name):
        """Fantasy team names occasionally carry a surname, e.g. "TEAM FOSTER"."""
        name = self.build_redactor()(name)
        if not name:
            return name
        for surname, display in TEAM_NAME_SURNAMES.items():
            name = re.sub(rf"(?<!\w){re.escape(surname)}(?!\w)", display, name, flags=re.IGNORECASE)
        return name

    def build_redactor(self):
        """
        Return shorten(text) -> text with every league member's full name replaced by
        their display name. Used on free text copied out of the workbook and by-laws
        (record books, draft notes, the rulebook) so no surname reaches the site.
        """
        if getattr(self, "_redactor", None):
            return self._redactor

        replacements: dict[str, str] = {}
        for full, display in self.full_to_display.items():
            first, last = full.split()[0], full.split()[-1]
            variants = {
                full,
                full.replace(" ", ""),
                f"{first}{last}",
                f"{first} {last}",
                f"{first}. {last}",
            }
            # "Matt Rose" is also written "Matthew Rose"; "Jon" as "John"; "Mike"/"Michael".
            for typo in TYPO_SURNAMES.get(last, []):
                variants.add(f"{first} {typo}")
                variants.add(typo if last in LONE_SURNAMES else f"{first} {typo}")
            for a, b in (("Matt", "Matthew"), ("Jon", "John"), ("Mike", "Michael"),
                         ("C.J.", "CJ"), ("Chris", "Christopher")):
                if first == a:
                    variants.add(f"{b} {last}")
                elif first == b:
                    variants.add(f"{a} {last}")
            if first == "C.J.":
                variants |= {f"Charles {last}", f"CJ {last}", f"C.J.{last}"}
            for variant in variants:
                replacements[variant.lower()] = display
        for surname, display in LONE_SURNAMES.items():
            replacements[surname.lower()] = display

        pattern = re.compile(
            r"(?<![\w.])(" + "|".join(re.escape(k) for k in sorted(replacements, key=len, reverse=True)) + r")(?!\w)",
            re.IGNORECASE,
        )

        def shorten(text):
            if not text or not isinstance(text, str):
                return text
            return pattern.sub(lambda m: replacements[m.group(1).lower()], text)

        self._redactor = shorten
        return shorten


# --------------------------------------------------------------------------------------
# Player name helpers
# --------------------------------------------------------------------------------------
# The lookbehind also excludes "." so initials like "J.K. Dobbins" keep their K.
_POS_TOKEN = re.compile(r"(?<![A-Za-z/.])(D/ST|DST|DEF|QB|RB|WR|TE|PK|K)(?![A-Za-z/.])")
_TRAILING_TEAM = re.compile(r"[,\s]+([A-Za-z]{2,4})\.?,?$")


def split_player(raw):
    """
    Roster cells come in several shapes across the years (and are riddled with
    non-breaking spaces and the odd stray token):
        "Brandon Marshall, NYJ WR"       (2014-2015)
        "Deebo Samuel Sr. Wsh, WR"       (2016+)
        "Eddie Lacy, GB\xa0RB\xa0\xa0K"  (typo in the source)
        "Dolphins D/ST Mia, D/ST"
        "Bijan Robinson"                 (draft boards / keepers)
    Strategy: cut the cell at the first standalone position token, then peel a trailing
    NFL team abbreviation off what's left. Returns (name, nfl_team, position).
    """
    if raw is None:
        return None, None, None
    text = re.sub(r"\s+", " ", str(raw).replace("\xa0", " ")).strip()
    if not text:
        return None, None, None

    nfl_team = pos = None
    m = _POS_TOKEN.search(text)
    if m and m.start() > 0:
        pos = m.group(1).upper()
        pos = {"DST": "D/ST", "DEF": "D/ST", "PK": "K"}.get(pos, pos)
        head = text[: m.start()].strip(" ,")
        # Team defenses read better keeping the D/ST in the name ("Dolphins D/ST").
        if pos == "D/ST":
            head = f"{head} D/ST"
        else:
            tm = _TRAILING_TEAM.search(head)
            if tm and norm_key(tm.group(1)) not in NAME_SUFFIXES:
                nfl_team = tm.group(1)
                head = head[: tm.start()].strip(" ,")
        text = head or text
    elif re.search(r"\bDefense\b", text, re.I):
        pos = "D/ST"

    text = text.strip(" ,")
    # ESPN questionable/injury tag glued onto the end: "Chris OlaveQ",
    # "Michael Pittman Jr.Q", "Kenneth Walker IIIQ". Leave all-caps strings alone.
    if len(text) > 2 and text.endswith("Q") and (text[-2].isalpha() or text[-2] == ".") and not text.isupper():
        text = text[:-1]
    # "Seahawks D" and "Seahawks D/ST" are the same defense.
    if re.search(r"\sD$", text):
        text = text[:-1] + "D/ST"
        pos = "D/ST"
    return text or None, nfl_team, pos


def player_key(name) -> str:
    """Match key for a player across sheets (handles Jr./Sr., punctuation, injury tags)."""
    display, _, _ = split_player(name)
    if not display:
        return ""
    # Defenses are written half a dozen ways ("Seahawks D", "Seattle Seahawks D/ST",
    # "New York Giants Defense"); key them on the nickname alone.
    dst = re.match(r"^(.*?)\s*(D/ST|DST|Defense)$", display, re.I)
    if dst:
        nickname = dst.group(1).split()[-1] if dst.group(1).split() else ""
        return norm_key(nickname) + "dst"
    key = norm_key(display)
    for suffix in sorted(NAME_SUFFIXES, key=len, reverse=True):
        if key.endswith(suffix) and len(key) > len(suffix) + 3:
            key = key[: -len(suffix)]
            break
    return key


# --------------------------------------------------------------------------------------
# Small sheet helpers
# --------------------------------------------------------------------------------------
def cell(ws, row, col):
    if row < 1 or col < 1:
        return None
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def as_text(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def as_int(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return int(round(float(v)))
        except (TypeError, ValueError):
            return None
    m = re.search(r"-?\d+", str(v))
    return int(m.group()) if m else None


def as_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    m = re.search(r"-?\d+(\.\d+)?", str(v).replace(",", ""))
    return round(float(m.group()), 4) if m else None


def sheet_year(name, suffix=None):
    """'2026 Draft' -> 2026 (with suffix='Draft'); '2025' -> 2025 (suffix=None)."""
    name = (name or "").strip()
    if suffix:
        m = re.fullmatch(rf"(\d{{4}})\s*{suffix}", name, re.I)
    else:
        m = re.fullmatch(r"(\d{4})", name)
    return int(m.group(1)) if m else None


def last_data_row(ws, cols, start=1, limit=None):
    limit = limit or ws.max_row
    last = start
    for r in range(start, limit + 1):
        if any(cell(ws, r, c) is not None for c in cols):
            last = r
    return last


# --------------------------------------------------------------------------------------
# Sheet parsers
# --------------------------------------------------------------------------------------
def find_sheet(wb, *candidates):
    """Sheet titles in this workbook carry stray trailing spaces; match loosely."""
    lookup = {norm_key(name): name for name in wb.sheetnames}
    for candidate in candidates:
        hit = lookup.get(norm_key(candidate))
        if hit:
            return wb[hit]
    return None


def parse_standings_history(wb, reg: OwnerRegistry):
    """
    'Standings History' lays each season out as a pair of columns
    (Regular Season | Playoff Final) with places 1..10 down the rows.
    Returns {year: {'regular': [owner...], 'playoff': [owner...]}} in finishing order.
    """
    ws = find_sheet(wb, "Standings History", "Standings History ")
    if ws is None:
        return {}

    header_row = None
    for r in range(1, 12):
        labels = [as_text(cell(ws, r, c)) for c in range(1, ws.max_column + 1)]
        if any(l and l.lower().startswith("regular season") for l in labels if l):
            header_row = r
            break
    if header_row is None:
        return {}

    year_row = header_row - 1
    seasons = {}
    for c in range(1, ws.max_column + 1):
        year = as_int(cell(ws, year_row, c))
        label = as_text(cell(ws, header_row, c)) or ""
        if year and 1900 < year < 2200 and label.lower().startswith("regular"):
            seasons[year] = {"reg_col": c, "final_col": c + 1}

    out = {}
    for year, cols in sorted(seasons.items()):
        regular, playoff = [], []
        for r in range(header_row + 1, header_row + 40):
            place = as_int(cell(ws, r, cols["reg_col"] - 1))
            reg_name = as_text(cell(ws, r, cols["reg_col"]))
            fin_name = as_text(cell(ws, r, cols["final_col"]))
            if not reg_name and not fin_name:
                if place is None:
                    break
                continue
            for raw, bucket in ((reg_name, regular), (fin_name, playoff)):
                if not raw:
                    continue
                resolved = reg.resolve(raw)
                bucket.append(
                    {
                        "place": place if place is not None else len(bucket) + 1,
                        "owner": resolved["display"] if resolved else raw,
                        "team": resolved["team"] if resolved else None,
                    }
                )
        if regular or playoff:
            out[year] = {"regular": regular, "playoff": playoff}
            for entry in regular + playoff:
                if entry["team"]:
                    reg.note_season(year, entry["owner"])
    return out


def parse_universal(wb, reg: OwnerRegistry):
    ws = find_sheet(wb, "Universal Standings")
    if ws is None:
        return {}

    def table(header_row, first_col, columns, key_col_offset=0):
        rows = []
        r = header_row + 1
        while r <= ws.max_row:
            raw_team = cell(ws, r, first_col)
            # Footnote rows below the table ("Team 4 previously | Dan Gordon") carry text
            # in the team-number column; a real row is numeric.
            if not isinstance(raw_team, (int, float)):
                break
            team = as_int(raw_team)
            owner_raw = as_text(cell(ws, r, first_col + 1))
            if team is None or not owner_raw:
                break
            resolved = reg.resolve(owner_raw)
            row = {
                "team": resolved["team"] if resolved else team,
                "owner": resolved["display"] if resolved else owner_raw,
            }
            for offset, (name, kind) in columns.items():
                v = cell(ws, r, first_col + offset)
                row[name] = as_int(v) if kind == "int" else as_float(v)
            rows.append(row)
            r += 1
        return rows

    # Locate the two header rows by their labels rather than fixed coordinates.
    points_header = finish_header = None
    points_col = finish_col = None
    for r in range(1, 20):
        for c in range(1, ws.max_column + 1):
            label = (as_text(cell(ws, r, c)) or "").lower()
            if label.startswith("total points") and points_header is None:
                points_header, points_col = r, c - 2
            if label.startswith("avg. combined") and finish_header is None:
                finish_header, finish_col = r, c - 2

    titles = []
    if points_header:
        titles = table(
            points_header,
            points_col,
            {
                2: ("points", "int"),
                3: ("championships", "int"),
                4: ("finalist", "int"),
                5: ("playoffs", "int"),
                6: ("trombley", "int"),
            },
        )
    finishes = []
    if finish_header:
        finishes = table(
            finish_header,
            finish_col,
            {
                2: ("avg_combined", "float"),
                3: ("avg_regular", "float"),
                4: ("avg_playoff", "float"),
            },
        )

    # Record blocks: "<something> Records" header, then Category | Record | Year | Owner.
    record_blocks = []
    for r in range(1, ws.max_row + 1):
        label = as_text(cell(ws, r, 2))
        if label and label.lower().endswith("records"):
            entries = []
            rr = r + 2  # skip the Category/Record/Year/Owner header line
            while rr <= ws.max_row:
                category = as_text(cell(ws, rr, 2))
                if not category:
                    break
                shorten = reg.build_redactor()
                entries.append(
                    {
                        "category": category,
                        "value": as_text(cell(ws, rr, 4)),
                        "year": as_text(cell(ws, rr, 5)),
                        # Shared records read "Matt Rose/Taylor Whitcomb" — free text.
                        "owner": shorten(as_text(cell(ws, rr, 6))),
                    }
                )
                rr += 1
            if entries:
                record_blocks.append({"title": label, "entries": entries})

    # Beer mile side table (label in one column, time in the next).
    beer_mile = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (as_text(cell(ws, r, c)) or "").lower() == "beer mile":
                rr = r + 1
                while rr <= ws.max_row:
                    who = as_text(cell(ws, rr, c))
                    time = as_text(cell(ws, rr, c + 1))
                    if not who or not time:
                        break
                    resolved = reg.resolve(who)
                    beer_mile.append({"owner": resolved["display"] if resolved else who, "time": time})
                    rr += 1
                break
        if beer_mile:
            break

    return {
        "titles": titles,
        "finishes": finishes,
        "records": record_blocks,
        "beer_mile": beer_mile,
    }


def parse_stats(wb, reg: OwnerRegistry):
    ws = find_sheet(wb, "Stats")
    if ws is None:
        return {"career": [], "seasons": {}}

    # ---- career tables (three side-by-side blocks sharing one header row) ----
    career = {}

    def career_block(header_row, first_col, columns):
        r = header_row + 1
        while r <= ws.max_row:
            raw_team = cell(ws, r, first_col)
            if not isinstance(raw_team, (int, float)):
                break
            team = as_int(raw_team)
            owner_raw = as_text(cell(ws, r, first_col + 1))
            if team is None or not owner_raw:
                break
            resolved = reg.resolve(owner_raw)
            key = resolved["team"] if resolved else team
            entry = career.setdefault(
                key,
                {"team": key, "owner": resolved["display"] if resolved else owner_raw},
            )
            for offset, (name, kind) in columns.items():
                v = cell(ws, r, first_col + offset)
                entry[name] = as_int(v) if kind == "int" else as_float(v)
            r += 1

    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            label = (as_text(cell(ws, r, c)) or "").lower()
            if label == "win":
                career_block(r, c - 2, {2: ("wins", "int"), 3: ("losses", "int"), 4: ("win_pct", "float")})
            elif label.startswith("points for"):
                career_block(
                    r,
                    c - 2,
                    {2: ("points_for", "float"), 3: ("points_against", "float"), 4: ("plus_minus", "float")},
                )
            elif label.startswith("acquires"):
                career_block(r, c - 2, {2: ("acquires", "int"), 3: ("trades", "int")})

    # ---- per-season blocks: a year in one column with "Team #" to its right ----
    seasons = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            year = as_int(cell(ws, r, c))
            nxt = (as_text(cell(ws, r, c + 1)) or "").lower()
            if year and 1990 < year < 2200 and nxt.startswith("team"):
                rows = []
                rr = r + 1
                while rr <= ws.max_row:
                    team = as_int(cell(ws, rr, c + 1))
                    owner_raw = as_text(cell(ws, rr, c + 2))
                    if team is None or not owner_raw:
                        break
                    resolved = reg.resolve(owner_raw, year)
                    rows.append(
                        {
                            "team": resolved["team"] if resolved else team,
                            "owner": resolved["display"] if resolved else owner_raw,
                            "wins": as_int(cell(ws, rr, c + 3)),
                            "losses": as_int(cell(ws, rr, c + 4)),
                            "points_for": as_float(cell(ws, rr, c + 5)),
                            "points_against": as_float(cell(ws, rr, c + 6)),
                            "acquires": as_int(cell(ws, rr, c + 7)),
                            "trades": as_int(cell(ws, rr, c + 8)),
                        }
                    )
                    rr += 1
                if rows:
                    seasons[year] = rows
                break

    return {"career": sorted(career.values(), key=lambda x: x["team"]), "seasons": seasons}


def parse_waivers(wb, reg: OwnerRegistry):
    ws = find_sheet(wb, "Waiver Wire Dollars")
    if ws is None:
        return {"years": [], "teams": [], "totals": {}, "note": None}

    header_row = None
    for r in range(1, 12):
        for c in range(1, ws.max_column + 1):
            if (as_text(cell(ws, r, c)) or "").lower().startswith("team"):
                header_row = r
                team_col = c
                break
        if header_row:
            break
    if header_row is None:
        return {"years": [], "teams": [], "totals": {}, "note": None}

    years = {}
    for c in range(team_col + 1, ws.max_column + 1):
        y = as_int(cell(ws, header_row, c))
        if y and 1990 < y < 2200:
            years[y] = c

    teams = []
    r = header_row + 1
    while r <= ws.max_row:
        team = as_int(cell(ws, r, team_col))
        owner_raw = as_text(cell(ws, r, team_col + 1))
        if team is None or not owner_raw:
            break
        resolved = reg.resolve(owner_raw)
        budgets = {}
        for y, c in years.items():
            v = as_float(cell(ws, r, c))
            if v is not None:
                budgets[str(y)] = v
        teams.append(
            {
                "team": resolved["team"] if resolved else team,
                "owner": resolved["display"] if resolved else owner_raw,
                "budgets": budgets,
            }
        )
        r += 1

    totals = {}
    note = None
    footer = r
    while footer <= min(ws.max_row, r + 3):
        label = as_text(cell(ws, footer, team_col)) or as_text(cell(ws, footer, team_col - 1))
        if label and "*" in label:
            note = label.lstrip("*").strip()
            for y, c in years.items():
                v = as_float(cell(ws, footer, c))
                if v is not None:
                    totals[str(y)] = v
            break
        footer += 1

    cap = None
    if note:
        m = re.search(r"\$(\d+)", note)
        if m:
            cap = int(m.group(1))

    return {
        "years": sorted(years.keys()),
        "teams": sorted(teams, key=lambda t: t["team"]),
        "totals": totals,
        "cap": cap,
        "note": reg.build_redactor()(note),
        "basis": "starting budget: $100 plus last season's leftover, capped",
    }


def find_keeper_sheet(wb):
    """
    The keeper tab is named for the season it feeds: "2026 Eligible Keepers". Older
    workbooks just called it "Keepers". Take the highest year present — that's the
    upcoming season — and fall back to the bare name.
    """
    years = []
    plain = None
    for name in wb.sheetnames:
        m = re.fullmatch(r"(\d{4})\s*(?:eligible\s*)?keepers", name.strip(), re.I)
        if m:
            years.append((int(m.group(1)), name))
        elif norm_key(name) == "keepers":
            plain = name
    if years:
        year, name = max(years)
        return year, wb[name]
    if plain:
        return None, wb[plain]
    return None, None


def parse_keepers(wb, reg: OwnerRegistry):
    year, ws = find_keeper_sheet(wb)
    if ws is None:
        return year, []

    header_row = header_col = None
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            if (as_text(cell(ws, r, c)) or "").lower() == "player":
                header_row, header_col = r, c
                break
        if header_row:
            break
    if header_row is None:
        return year, []

    # Map the header labels so a re-ordered/extended sheet still parses.
    fields = {}
    for c in range(header_col, ws.max_column + 1):
        label = (as_text(cell(ws, header_row, c)) or "").lower().rstrip("?").strip()
        if not label:
            continue
        if label == "player":
            fields["player"] = c
        elif label.startswith("team"):
            fields["team"] = c
        elif "owner" in label:
            fields["owner"] = c
        elif label.startswith("year signed"):
            fields["year_signed"] = c
        elif label.startswith("rookie"):
            fields["rookie"] = c
        elif label.startswith("contract"):
            fields["contract_years"] = c
        elif label.startswith("acq"):
            fields["acquired"] = c
        elif "keeper round" in label:
            fields["cost"] = c

    rows = []
    r = header_row + 1
    blanks = 0
    while r <= ws.max_row and blanks < 5:
        raw_player = as_text(cell(ws, r, fields.get("player", header_col)))
        if not raw_player:
            blanks += 1
            r += 1
            continue
        blanks = 0
        name, nfl_team, pos = split_player(raw_player)
        team = as_int(cell(ws, r, fields["team"])) if "team" in fields else None
        owner_raw = as_text(cell(ws, r, fields["owner"])) if "owner" in fields else None
        resolved = reg.resolve(owner_raw) if owner_raw else None
        cost_raw = cell(ws, r, fields["cost"]) if "cost" in fields else None
        cost_round = as_int(cost_raw)
        cost_text = as_text(cost_raw)
        rows.append(
            {
                "player": name,
                "player_key": player_key(raw_player),
                "nfl_team": nfl_team,
                "position": pos,
                "team": resolved["team"] if resolved else team,
                "owner": resolved["display"] if resolved else owner_raw,
                "year_signed": as_int(cell(ws, r, fields["year_signed"])) if "year_signed" in fields else None,
                "rookie": (as_text(cell(ws, r, fields["rookie"])) or "").lower().startswith("y")
                if "rookie" in fields
                else False,
                "contract_years_remaining": as_int(cell(ws, r, fields["contract_years"]))
                if "contract_years" in fields
                else None,
                "acquired": as_text(cell(ws, r, fields["acquired"])) if "acquired" in fields else None,
                "cost_round": cost_round if cost_text and cost_text.isdigit() else None,
                "cost_label": cost_text,
                "eligible": bool(cost_text) and not (cost_text or "").lower().startswith("return"),
            }
        )
        r += 1
    return year, rows


def parse_draft_board(ws, year, reg: OwnerRegistry):
    """
    Draft board layout: row 1 is a title, row 2 holds owner first names across the
    columns (left to right = draft order), then one row per round, then a block of
    keeper / trade annotations underneath.
    """
    title = None
    for c in range(1, ws.max_column + 1):
        t = as_text(cell(ws, 1, c))
        if t:
            title = t
            break

    header_row = None
    for r in range(1, 8):
        labels = [as_text(cell(ws, r, c)) for c in range(1, ws.max_column + 1)]
        if sum(1 for l in labels if l) >= 6 and not any(
            (l or "").lower().startswith("round") for l in labels
        ):
            header_row = r
            break
    if header_row is None:
        header_row = 2

    columns = []
    for c in range(1, ws.max_column + 1):
        raw = as_text(cell(ws, header_row, c))
        if not raw or raw.lower().startswith("round"):
            continue
        resolved = reg.resolve(raw, year)
        columns.append(
            {
                "col": c,
                "slot": len(columns) + 1,
                "label": raw,
                "owner": resolved["display"] if resolved else raw,
                "team": resolved["team"] if resolved else None,
            }
        )

    label_col = min((c["col"] for c in columns), default=3) - 1
    rounds = []
    notes = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = as_text(cell(ws, r, label_col))
        values = {c["slot"]: as_text(cell(ws, r, c["col"])) for c in columns}
        if not label and not any(values.values()):
            continue
        m = re.fullmatch(r"round\s*(\d+)", (label or "").lower())
        if m:
            picks = []
            for col in columns:
                raw = values.get(col["slot"])
                traded = None
                if raw and re.match(r"^\s*trade[d]?\s+(to|w/?|with)\b", raw, re.I):
                    traded = re.sub(r"^\s*trade[d]?\s+(to|w/?|with)\s*", "", raw, flags=re.I).strip()
                    raw = None
                name, nfl_team, pos = split_player(raw)
                picks.append(
                    {
                        "slot": col["slot"],
                        "owner": col["owner"],
                        "team": col["team"],
                        "player": name,
                        "player_key": player_key(raw) if name else "",
                        "nfl_team": nfl_team,
                        "position": pos,
                        "traded_to": traded,
                    }
                )
            rounds.append({"round": int(m.group(1)), "picks": picks})
        elif any(values.values()) or label:
            shorten = reg.build_redactor()
            entries = [
                # Note cells read like "Round 3 via Chris (open)" — free text.
                {"owner": col["owner"], "team": col["team"], "text": shorten(values.get(col["slot"]))}
                for col in columns
                if values.get(col["slot"])
            ]
            if entries:
                notes.append({"label": label or "Note", "entries": entries})

    return {
        "year": year,
        "title": title,
        "source": "draft board",
        # `label` is what the board column is titled; use the resolved display name so a
        # raw header like "Aberdale" or "Caruso" never reaches the site.
        "order": [{"slot": c["slot"], "owner": c["owner"], "team": c["team"], "label": c["owner"]} for c in columns],
        "rounds": rounds,
        "notes": notes,
    }


def parse_roster_sheet(ws, year, reg: OwnerRegistry):
    """
    Roster sheets stack repeating 3-column blocks. A draft block is anchored by a
    "Keeper?" label, a final-roster block by "ACG?" (acquisition). For each anchor the
    owner name sits one row above in the first column of the block, and the fantasy team
    name on the anchor row itself.
    """
    anchors = {"Keeper?": [], "ACG?": []}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = as_text(cell(ws, r, c))
            if v in anchors:
                anchors[v].append((r, c))

    def read_block(anchor_row, anchor_col):
        first = anchor_col - 2
        owner_raw = as_text(cell(ws, anchor_row - 1, first))
        team_name = as_text(cell(ws, anchor_row, first))
        # One block in the 2022 sheet has a team name typed where the owner belongs;
        # leave it unresolved here and infer the franchise afterwards.
        resolved = reg.resolve(owner_raw, year, record=False) if owner_raw else None
        entries = []
        r = anchor_row + 1
        blanks = 0
        while r <= ws.max_row and blanks < 2:
            left = cell(ws, r, first)
            mid = as_text(cell(ws, r, first + 1))
            flag = as_text(cell(ws, r, first + 2))
            if left is None and mid is None:
                blanks += 1
                r += 1
                continue
            # Stop if we've run into the next block's owner header.
            if mid is None and flag is None and isinstance(left, str) and reg.resolve(left, year, record=False):
                break
            blanks = 0
            name, nfl_team, pos = split_player(mid)
            entries.append(
                {
                    "left": as_text(left),
                    "left_num": as_int(left),
                    "player": name,
                    "player_key": player_key(mid) if name else "",
                    "nfl_team": nfl_team,
                    "position": pos,
                    "flag": flag,
                }
            )
            r += 1
        return {
            "owner": resolved["display"] if resolved else owner_raw,
            "team": resolved["team"] if resolved else None,
            "team_name": team_name,
            "entries": entries,
        }

    drafts = [read_block(r, c) for r, c in anchors["Keeper?"]]
    rosters = [read_block(r, c) for r, c in anchors["ACG?"]]

    def infer_missing_owners(blocks):
        """A block whose owner cell is unusable gets the one franchise nobody claimed."""
        known = {b["team"] for b in blocks if b["team"]}
        orphans = [b for b in blocks if not b["team"]]
        missing = [t for t in sorted(reg.teams) if t not in known]
        if len(orphans) == 1 and len(missing) == 1:
            orphans[0]["team"] = missing[0]
            orphans[0]["owner"] = reg.owner_for_team(missing[0])
        else:
            for block in orphans:
                if block["owner"]:
                    reg.unresolved[block["owner"]] += 1

    infer_missing_owners(rosters)
    infer_missing_owners(drafts)

    # The pick column is the round in 2016+ sheets but an overall pick number in the
    # earliest ones; normalise to a round using the number of franchises in the league.
    team_count = max(len(drafts), 1)
    max_pick = max(
        (e["left_num"] or 0 for block in drafts for e in block["entries"]),
        default=0,
    )
    overall_numbering = max_pick > team_count * 3

    draft_blocks = []
    for block in drafts:
        picks = []
        for entry in block["entries"]:
            num = entry["left_num"]
            if num is None:
                continue
            rnd = math.ceil(num / team_count) if overall_numbering else num
            picks.append(
                {
                    "round": rnd,
                    "overall": num if overall_numbering else None,
                    "player": entry["player"],
                    "player_key": entry["player_key"],
                    "nfl_team": entry["nfl_team"],
                    "position": entry["position"],
                    "keeper": (entry["flag"] or "").lower().startswith("y"),
                }
            )
        draft_blocks.append(
            {
                "owner": block["owner"],
                "team": block["team"],
                "team_name": reg.shorten_team_name(block["team_name"]),
                "picks": picks,
            }
        )

    roster_blocks = []
    for block in rosters:
        record = None
        name = block["team_name"]
        if name:
            m = re.search(r"\((\d+)\s*-\s*(\d+)\)", name)
            if m:
                record = f"{m.group(1)}-{m.group(2)}"
                name = name[: m.start()].strip()
        roster_blocks.append(
            {
                "owner": block["owner"],
                "team": block["team"],
                "team_name": reg.shorten_team_name(name),
                "record": record,
                "players": [
                    {
                        "slot": e["left"],
                        "player": e["player"],
                        "player_key": e["player_key"],
                        "nfl_team": e["nfl_team"],
                        "position": e["position"],
                        "acquired": e["flag"],
                    }
                    for e in block["entries"]
                    if e["player"]
                ],
            }
        )

    # Left-hand "Player | Team # | Round" index present on the newer roster sheets.
    index_rows = []
    header_row = header_col = None
    for r in range(1, 10):
        for c in range(1, 6):
            if (as_text(cell(ws, r, c)) or "").lower() == "player" and (
                (as_text(cell(ws, r, c + 1)) or "").lower().startswith("team")
            ):
                header_row, header_col = r, c
                break
        if header_row:
            break
    if header_row:
        r = header_row + 1
        blanks = 0
        while r <= ws.max_row and blanks < 4:
            raw = as_text(cell(ws, r, header_col))
            if not raw:
                blanks += 1
                r += 1
                continue
            blanks = 0
            name, nfl_team, pos = split_player(raw)
            team = as_int(cell(ws, r, header_col + 1))
            index_rows.append(
                {
                    "player": name,
                    "player_key": player_key(raw),
                    "nfl_team": nfl_team,
                    "position": pos,
                    "team": team,
                    "owner": reg.owner_for_team(team),
                    "round": as_int(cell(ws, r, header_col + 2)),
                }
            )
            r += 1

    return {"year": year, "drafts": draft_blocks, "rosters": roster_blocks, "draft_index": index_rows}


def board_from_roster(year, roster_data):
    """Synthesise a browsable draft board for seasons with no dedicated draft sheet."""
    blocks = [b for b in roster_data["drafts"] if b["picks"]]
    if not blocks:
        return None
    order = []
    for slot, block in enumerate(blocks, start=1):
        order.append({"slot": slot, "owner": block["owner"], "team": block["team"], "label": block["owner"]})
    max_round = max((p["round"] for b in blocks for p in b["picks"] if p["round"]), default=0)
    rounds = []
    for rnd in range(1, max_round + 1):
        picks = []
        for slot, block in enumerate(blocks, start=1):
            hit = next((p for p in block["picks"] if p["round"] == rnd), None)
            picks.append(
                {
                    "slot": slot,
                    "owner": block["owner"],
                    "team": block["team"],
                    "player": hit["player"] if hit else None,
                    "player_key": hit["player_key"] if hit else "",
                    "nfl_team": hit["nfl_team"] if hit else None,
                    "position": hit["position"] if hit else None,
                    "keeper": bool(hit and hit["keeper"]),
                    "raw": hit["player"] if hit else None,
                }
            )
        rounds.append({"round": rnd, "picks": picks})
    return {
        "year": year,
        "title": f"{year} Draft",
        "source": "roster sheet",
        "order": order,
        "rounds": rounds,
        "notes": [],
    }


# --------------------------------------------------------------------------------------
# By-laws (.docx)
# --------------------------------------------------------------------------------------
SECTION_RE = re.compile(r"^\[?\s*(\d+(?:\.\d+)*)\s+(.+?)\s*\]?$")

# The site is public, so email addresses in the by-laws are replaced on the way out.
# The source .docx is never modified. Pass --keep-emails to publish them verbatim.
EMAIL_RE = re.compile(r"\b[\w.+-]+@[\w-]+\.[\w.-]+\b")
EMAIL_PLACEHOLDER = "[email — ask the commissioner]"


def esc(text):
    return (
        (text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def parse_bylaws(path, img_dir, img_rel, shorten=None, keep_emails=False):
    if python_docx is None:
        print("  ! python-docx not installed; skipping rulebook", file=sys.stderr)
        return None
    if not path or not os.path.exists(path):
        print("  ! by-laws document not found; skipping rulebook", file=sys.stderr)
        return None

    doc = python_docx.Document(path)
    os.makedirs(img_dir, exist_ok=True)

    # Map relationship ids to extracted image files so pictures land in the right place.
    saved = {}

    def save_image(rid):
        if rid in saved:
            return saved[rid]
        part = doc.part.related_parts[rid]
        blob = part.blob
        ext = os.path.splitext(part.partname)[1] or ".png"
        digest = hashlib.sha1(blob).hexdigest()[:10]
        fname = f"bylaws-{digest}{ext}"
        with open(os.path.join(img_dir, fname), "wb") as fh:
            fh.write(blob)
        saved[rid] = f"{img_rel}/{fname}"
        return saved[rid]

    sections = []
    current = None
    intro = []

    def para_images(p):
        rids = re.findall(r'r:embed="([^"]+)"', p._p.xml)
        return [save_image(rid) for rid in rids]

    def list_level(p):
        pr = p._p.pPr
        if pr is not None and pr.numPr is not None and pr.numPr.ilvl is not None:
            return int(pr.numPr.ilvl.val)
        if p.style.name == "List Paragraph":
            return 0
        return None

    redacted = 0
    for p in doc.paragraphs:
        text = re.sub(r"\s+", " ", p.text or "").strip()
        if not keep_emails and EMAIL_RE.search(text):
            text, hits = EMAIL_RE.subn(EMAIL_PLACEHOLDER, text)
            redacted += hits
        if shorten:
            text = shorten(text)
        images = para_images(p)
        m = SECTION_RE.match(text) if text.startswith("[") else None
        if m:
            current = {
                "number": m.group(1),
                "title": m.group(2).strip(),
                "level": m.group(1).count(".") + 1,
                "blocks": [],
            }
            sections.append(current)
            continue
        target = current["blocks"] if current else intro
        for src in images:
            target.append({"type": "image", "src": src})
        if not text:
            continue
        level = list_level(p)
        if level is not None:
            target.append({"type": "li", "level": level, "text": text})
        else:
            target.append({"type": "p", "text": text})

    def blocks_html(blocks):
        html = []
        open_levels = []
        for block in blocks:
            if block["type"] == "li":
                level = block["level"]
                while open_levels and open_levels[-1] > level:
                    html.append("</ul>")
                    open_levels.pop()
                if not open_levels or open_levels[-1] < level:
                    html.append('<ul class="rb-list">')
                    open_levels.append(level)
                html.append(f"<li>{esc(block['text'])}</li>")
            else:
                while open_levels:
                    html.append("</ul>")
                    open_levels.pop()
                if block["type"] == "image":
                    html.append(
                        f'<figure class="rb-figure"><img src="{block["src"]}" alt="By-laws exhibit" loading="lazy"></figure>'
                    )
                else:
                    html.append(f"<p>{esc(block['text'])}</p>")
        while open_levels:
            html.append("</ul>")
            open_levels.pop()
        return "\n".join(html)

    out_sections = []
    for s in sections:
        out_sections.append(
            {
                "number": s["number"],
                "title": s["title"],
                "level": s["level"],
                "html": blocks_html(s["blocks"]),
                "text": " ".join(b.get("text", "") for b in s["blocks"] if b.get("text")),
            }
        )

    if redacted:
        print(f"  rulebook: {redacted} email address(es) redacted")

    return {
        "source": os.path.basename(path),
        "intro_html": blocks_html(intro),
        "sections": out_sections,
    }


# --------------------------------------------------------------------------------------
# Overlays
# --------------------------------------------------------------------------------------
# A season sometimes lands before it has been typed into the workbook (results come off
# ESPN first). Files in source/ fill those gaps. The workbook always wins: an overlay is
# ignored the moment the same year shows up in Standings History / Stats / Waiver Wire.
def load_overlay(name):
    path = os.path.join(ROOT, "source", name)
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def overlay_seasons():
    """Yield every source/season-<year>.json, oldest first."""
    hits = []
    for path in sorted(glob.glob(os.path.join(ROOT, "source", "season-*.json"))):
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        if data.get("year"):
            hits.append(data)
    return sorted(hits, key=lambda d: d["year"])


def apply_season_overlay(data, reg: OwnerRegistry, standings, stats, waivers):
    """Merge one season overlay into the parsed workbook data. Returns True if applied."""
    year = data["year"]
    applied = False

    if year not in standings and (data.get("regular_season") or data.get("playoff_final")):
        standings[year] = {
            "regular": [
                {"place": e["place"], "owner": reg.owner_for_team(e["team"]), "team": e["team"]}
                for e in sorted(data.get("regular_season", []), key=lambda e: e["place"])
            ],
            "playoff": [
                {"place": e["place"], "owner": reg.owner_for_team(e["team"]), "team": e["team"]}
                for e in sorted(data.get("playoff_final", []), key=lambda e: e["place"])
            ],
        }
        for entry in standings[year]["regular"] + standings[year]["playoff"]:
            if entry["owner"]:
                reg.note_season(year, entry["owner"])
        applied = True

    if year not in stats["seasons"] and data.get("stats"):
        stats["seasons"][year] = [
            {
                "team": row["team"],
                "owner": reg.owner_for_team(row["team"]),
                "wins": row.get("wins"),
                "losses": row.get("losses"),
                "ties": row.get("ties") or None,
                "points_for": row.get("points_for"),
                "points_against": row.get("points_against"),
                "acquires": row.get("acquires"),
                "trades": row.get("trades"),
            }
            for row in data["stats"]
        ]
        applied = True

    remaining = data.get("waiver_remaining") or data.get("waiver_budgets") or {}
    if remaining:
        waivers.setdefault("remaining", {})[str(year)] = {
            str(team): value for team, value in remaining.items()
        }
        applied = True

    return applied


def roll_waivers_forward(waivers):
    """
    The workbook's yearly columns are starting budgets: $100 plus whatever was left at the
    end of the season before, capped (by-laws 6.3).

    Wherever we have end-of-season numbers, compute the next season's start from them
    rather than reading the sheet. That matters for the tail of the series: the workbook's
    last column is headed 2024 but its values are exactly $100 + the end-of-2024 leftovers,
    so it is really the 2025 starting budget, and the true 2024 column was never recorded.
    """
    overlay = load_overlay("waiver-end-of-season.json") or {}
    remaining = dict(overlay.get("remaining") or {})
    remaining.update(waivers.get("remaining") or {})
    if not remaining:
        return
    waivers["remaining"] = remaining

    caps = overlay.get("cap_by_year") or {}
    default_cap = caps.get("default") or waivers.get("cap") or 150

    computed = []
    for year_str, leftovers in sorted(remaining.items()):
        season = int(year_str) + 1
        cap = caps.get(str(season), default_cap)
        for team in waivers["teams"]:
            left = leftovers.get(str(team["team"]))
            if left is not None:
                team["budgets"][str(season)] = min(cap, 100 + left)
        computed.append(season)

    # The computed values simply overwrite the sheet's last (mislabelled) column, so
    # nothing needs removing — 2023 and earlier are read from the workbook as-is.
    waivers["years"] = sorted(set(waivers["years"]) | set(computed))
    waivers["computed_years"] = sorted(computed)
    waivers["cap_by_year"] = caps
    waivers["note"] = (
        "Starting budget is $100 plus what was left at the end of the season before, "
        f"capped at ${caps.get(str(max(computed)), default_cap)}."
    )


def augment_career(stats, universal, reg: OwnerRegistry, overlay_years):
    """
    Add an overlay season to the workbook's career totals rather than recomputing them.
    The commissioner keeps those columns (and the universal points table) by hand, and a
    from-scratch recount quietly disagrees with them on old seasons — so only ever apply
    the delta for the season the workbook doesn't carry yet.
    """
    career = {row["team"]: row for row in stats["career"]}
    for year in overlay_years:
        for row in stats["seasons"].get(year, []):
            entry = career.setdefault(
                row["team"],
                {"team": row["team"], "owner": reg.owner_for_team(row["team"]),
                 "wins": 0, "losses": 0, "points_for": 0.0, "points_against": 0.0,
                 "acquires": 0, "trades": 0},
            )
            for key in ("wins", "losses", "ties", "acquires", "trades"):
                if row.get(key):
                    entry[key] = (entry.get(key) or 0) + row[key]
            for key in ("points_for", "points_against"):
                if row.get(key):
                    entry[key] = round((entry.get(key) or 0.0) + row[key], 2)

    for entry in career.values():
        games = (entry.get("wins") or 0) + (entry.get("losses") or 0) + (entry.get("ties") or 0)
        if games:
            entry["win_pct"] = round(((entry.get("wins") or 0) + (entry.get("ties") or 0) / 2) / games, 4)
        if entry.get("points_for") is not None and entry.get("points_against") is not None:
            entry["plus_minus"] = round(entry["points_for"] - entry["points_against"], 2)
    stats["career"] = sorted(career.values(), key=lambda e: e["team"])


def augment_universal(universal, standings, reg: OwnerRegistry, overlay_years):
    """
    Same idea for the universal table: championship 3, finals appearance 2, playoff berth
    1, the Moules -1 — applied only to the overlay seasons, on top of the stored totals.
    """
    titles = {row["team"]: row for row in universal.get("titles", [])}
    for year in overlay_years:
        for entry in standings.get(year, {}).get("playoff", []):
            team, place = entry["team"], entry["place"]
            if not team:
                continue
            row = titles.setdefault(
                team,
                {"team": team, "owner": reg.owner_for_team(team), "points": 0,
                 "championships": 0, "finalist": 0, "playoffs": 0, "trombley": 0},
            )
            if place == 1:
                row["championships"] += 1
                row["points"] += 3
            elif place == 2:
                row["finalist"] += 1
                row["points"] += 2
            if place <= 6:
                row["playoffs"] += 1
                row["points"] += 1
            if place == 10:
                row["trombley"] += 1
                row["points"] -= 1
    universal["titles"] = sorted(titles.values(), key=lambda r: (-(r["points"] or 0), r["team"]))

    # Average finishes are a straight mean over every season on record, so those can be
    # recomputed outright.
    finishes = {}
    for season in standings.values():
        for bucket in ("regular", "playoff"):
            for entry in season.get(bucket) or []:
                if entry["team"]:
                    finishes.setdefault(entry["team"], {"regular": [], "playoff": []})[bucket].append(entry["place"])
    universal["finishes"] = sorted(
        (
            {
                "team": team,
                "owner": reg.owner_for_team(team),
                "avg_regular": round(sum(v["regular"]) / len(v["regular"]), 4) if v["regular"] else None,
                "avg_playoff": round(sum(v["playoff"]) / len(v["playoff"]), 4) if v["playoff"] else None,
                "avg_combined": round(
                    (sum(v["regular"]) + sum(v["playoff"])) / (len(v["regular"]) + len(v["playoff"])), 4
                )
                if (v["regular"] or v["playoff"])
                else None,
            }
            for team, v in finishes.items()
        ),
        key=lambda r: (r["avg_combined"] if r["avg_combined"] is not None else 99),
    )


# --------------------------------------------------------------------------------------
# Derived data
# --------------------------------------------------------------------------------------
def build_player_index(boards, roster_years, keepers):
    """
    player_key -> {name, history:[{year, round, owner, team, keeper}]}

    The draft board is authoritative for the round a player went in: owners who traded
    picks have roster-sheet pick numbers that count their own selections rather than the
    true round. Roster sheets therefore only contribute keeper flags for seasons that
    have a real draft board, and full picks for the seasons that don't.
    """
    index = {}

    def add(key, name, entry):
        if not key or not name:
            return
        rec = index.setdefault(key, {"key": key, "name": name, "history": []})
        # Prefer the longest spelling we've seen as the display name.
        if len(name) > len(rec["name"]):
            rec["name"] = name
        for existing in rec["history"]:
            if existing["year"] != entry["year"]:
                continue
            # Same season + same owner (or same round) is the same pick seen twice.
            if existing["owner"] == entry["owner"] or existing["round"] == entry["round"]:
                for k, v in entry.items():
                    if k == "keeper":
                        existing["keeper"] = existing.get("keeper") or v
                    elif v is not None:
                        existing.setdefault(k, v)
                return
        rec["history"].append(entry)

    for board in boards:
        for rnd in board["rounds"]:
            for pick in rnd["picks"]:
                if not pick["player"]:
                    continue
                add(
                    pick["player_key"],
                    pick["player"],
                    {
                        "year": board["year"],
                        "round": rnd["round"],
                        "owner": pick["owner"],
                        "team": pick["team"],
                        "keeper": bool(pick.get("keeper")),
                        "position": pick.get("position"),
                        "source": "board",
                    },
                )

    for year, data in roster_years.items():
        for block in data["drafts"]:
            for pick in block["picks"]:
                if not pick["player"]:
                    continue
                add(
                    pick["player_key"],
                    pick["player"],
                    {
                        "year": year,
                        "round": pick["round"],
                        "owner": block["owner"],
                        "team": block["team"],
                        "keeper": pick["keeper"],
                        "position": pick.get("position"),
                        "source": "roster",
                    },
                )

    for row in keepers:
        if row["player"]:
            index.setdefault(row["player_key"], {"key": row["player_key"], "name": row["player"], "history": []})

    for rec in index.values():
        rec["history"].sort(key=lambda h: (h["year"], h["round"] or 99))
    return index


def build_draft_prep(rankings, keepers, year):
    """
    Join ESPN's preseason ranking to the keeper table so a draft board can be read as
    "what would this player cost me, and is he worth it".

    ESPN's own rank numbers have holes in them (nothing is ranked 37-68 this year, and ADP
    runs straight through the gap), so the published rank is the position in their ordered
    list — ESPN's #1, #2, #3 — with their raw value kept alongside.

    Value = keeper cost weighted at ten per round, minus the ranking. A second-round
    keeper is charged 20, so a top-five player kept for a 2nd scores +15. Higher is better.
    """
    if not rankings:
        return None

    by_key = {}
    for row in keepers:
        if row.get("player_key"):
            by_key.setdefault(row["player_key"], row)

    # The league plays half-point PPR and ESPN publishes no half-PPR list, so blend their
    # standard and full-PPR ranks — the two scoring systems this sits exactly between —
    # and re-rank off the average. Players ESPN ranks in only one system keep that rank.
    ordered = sorted(
        rankings["players"],
        key=lambda p: ((p["rank"] + (p.get("rank_standard") or p["rank"])) / 2, p["rank"]),
    )

    rows = []
    matched = 0
    for ordinal, entry in enumerate(ordered, start=1):
        key = player_key(entry["player"])
        keeper = by_key.get(key)
        if not keeper:
            close = difflib.get_close_matches(key, by_key.keys(), n=1, cutoff=0.9)
            keeper = by_key[close[0]] if close else None
        if keeper:
            matched += 1
        cost = keeper["cost_round"] if keeper else None
        rows.append(
            {
                "rank": ordinal,
                "rank_ppr": entry["rank"],
                "rank_standard": entry.get("rank_standard"),
                "player": entry["player"],
                "player_key": key,
                "nfl_team": entry["nfl_team"],
                "position": entry["position"],
                "adp": entry.get("adp"),
                "owner": keeper["owner"] if keeper else None,
                "team": keeper["team"] if keeper else None,
                "cost_round": cost,
                "cost_label": keeper["cost_label"] if keeper else None,
                "years_remaining": keeper["contract_years_remaining"] if keeper else None,
                "value": (cost * 10 - ordinal) if cost else None,
            }
        )

    return {
        "year": year,
        "source": rankings.get("source"),
        "scoring": "half-PPR blend of ESPN's standard and full-PPR ranks",
        "weighting": "keeper cost counts as ten points per round; value = (round x 10) - ranking",
        "matched_to_rosters": matched,
        "players": rows,
    }


def draft_order_from_standings(final_places):
    """
    By-laws 4.2: the right to pick a draft slot is handed out in this order of finish.
    final_places maps place -> owner for the most recent completed season.
    """
    preference = [
        (7, "Winner of the Consolation Bracket"),
        (8, "Runner-up of the Consolation Bracket"),
        (9, "Winner of the Moules/Trombley Bowl"),
        (10, "The Moules (last place)"),
        (5, "Fifth place"),
        (6, "Sixth place"),
        (3, "Third place"),
        (4, "Fourth place"),
        (2, "Runner-up"),
        (1, "League Champion"),
    ]
    order = []
    for choice, (place, label) in enumerate(preference, start=1):
        owner = final_places.get(place)
        if owner:
            order.append({"choice": choice, "place": place, "label": label, "owner": owner})
    return order


# --------------------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------------------
def discover(pattern, *dirs):
    for d in dirs:
        hits = sorted(glob.glob(os.path.join(d, pattern)))
        hits = [h for h in hits if not os.path.basename(h).startswith("~$")]
        if hits:
            return hits[-1]
    return None


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
    return os.path.getsize(path)


def main():
    ap = argparse.ArgumentParser(description="Build O8FFL site data from the master workbook.")
    ap.add_argument("--workbook", help="path to the .xlsx master workbook")
    ap.add_argument("--bylaws", help="path to the by-laws .docx")
    ap.add_argument("--out", default=os.path.join(ROOT, "data"), help="output directory for JSON")
    ap.add_argument("--assets", default=os.path.join(ROOT, "assets"), help="site assets directory")
    ap.add_argument(
        "--keep-emails",
        action="store_true",
        help="publish email addresses from the by-laws verbatim (default: redact them)",
    )
    args = ap.parse_args()

    source_dirs = [os.path.join(ROOT, "source"), ROOT]
    workbook_path = args.workbook or discover("*.xlsx", *source_dirs)
    bylaws_path = args.bylaws or discover("*By-Laws*.docx", *source_dirs) or discover("*.docx", *source_dirs)

    if not workbook_path or not os.path.exists(workbook_path):
        sys.exit("Could not find the master workbook (.xlsx). Pass --workbook.")

    print(f"workbook : {os.path.relpath(workbook_path, ROOT)}")
    print(f"by-laws  : {os.path.relpath(bylaws_path, ROOT) if bylaws_path else '(none)'}")

    os.makedirs(args.out, exist_ok=True)
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    reg = OwnerRegistry()

    # Seasons first: they teach the registry who played when, which the draft-board
    # first-name headers need in order to disambiguate.
    standings = parse_standings_history(wb, reg)
    print(f"  standings history: {len(standings)} seasons")

    universal = parse_universal(wb, reg)
    stats = parse_stats(wb, reg)
    waivers = parse_waivers(wb, reg)
    keeper_year, keepers = parse_keepers(wb, reg)
    print(
        f"  universal: {len(universal.get('titles', []))} teams | "
        f"stats: {len(stats['seasons'])} seasons | waivers: {len(waivers['years'])} years | "
        f"keepers: {len(keepers)} rows"
    )

    # Roster sheets (bare year names) and draft boards ("<year> Draft").
    roster_years = {}
    for name in wb.sheetnames:
        year = sheet_year(name)
        if year:
            roster_years[year] = parse_roster_sheet(wb[name], year, reg)
    print(f"  roster sheets: {sorted(roster_years)}")

    boards = {}
    for name in wb.sheetnames:
        year = sheet_year(name, "Draft")
        if year:
            boards[year] = parse_draft_board(wb[name], year, reg)
    for year, data in roster_years.items():
        if year not in boards:
            synth = board_from_roster(year, data)
            if synth:
                boards[year] = synth

    # Seasons with no roster sheet have no keeper column, but build_keepers.py works out
    # who was kept from where they were taken relative to their cost — fold that in too.
    reconstructed = {}
    for path in glob.glob(os.path.join(ROOT, "source", "eligible-keepers-*.json")):
        with open(path, encoding="utf-8") as fh:
            doc = json.load(fh)
        for entry in doc.get("prior_keepers") or []:
            reconstructed.setdefault(doc["year"] - 1, set()).add((entry["team"], entry["player_key"]))

    # Fold keeper flags from the roster sheets into the boards.
    for year, board in boards.items():
        roster = roster_years.get(year)
        keeper_keys = set(reconstructed.get(year, ()))
        if not roster:
            if keeper_keys:
                for rnd in board["rounds"]:
                    for pick in rnd["picks"]:
                        pick["keeper"] = bool(
                            pick["player_key"] and (pick["team"], pick["player_key"]) in keeper_keys
                        )
            continue
        keeper_keys |= {
            (block["team"], p["player_key"])
            for block in roster["drafts"]
            for p in block["picks"]
            if p["keeper"] and p["player_key"]
        }
        for rnd in board["rounds"]:
            for pick in rnd["picks"]:
                if pick["player_key"] and (pick["team"], pick["player_key"]) in keeper_keys:
                    pick["keeper"] = True
                else:
                    pick.setdefault("keeper", False)
    print(f"  draft boards: {sorted(boards)}")

    # ---- overlays for seasons the workbook hasn't absorbed yet -------------------------
    merged = []
    for overlay in overlay_seasons():
        if apply_season_overlay(overlay, reg, standings, stats, waivers):
            merged.append(overlay["year"])
    roll_waivers_forward(waivers)

    if merged:
        augment_career(stats, universal, reg, merged)
        augment_universal(universal, standings, reg, merged)
        print(f"  overlay seasons merged: {merged} (career + universal totals extended)")

    # Final rosters that only exist as an overlay (screenshots, not a workbook tab).
    for path in sorted(glob.glob(os.path.join(ROOT, "source", "final-rosters-*.json"))):
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        year = data.get("year")
        if not year or year in roster_years:
            continue
        roster_years[year] = {
            "year": year,
            "drafts": [],
            "rosters": [
                {
                    "owner": reg.owner_for_team(block["team"]),
                    "team": block["team"],
                    "team_name": reg.shorten_team_name(block.get("team_name")),
                    "record": None,
                    "players": [
                        {
                            "slot": slot,
                            "player": name,
                            "player_key": player_key(name),
                            "nfl_team": nfl_team,
                            "position": pos,
                            "acquired": acq,
                        }
                        for slot, name, nfl_team, pos, acq in block["players"]
                    ],
                }
                for block in data["teams"]
            ],
            "draft_index": [],
        }
        print(f"  final rosters overlay: {year} ({len(data['teams'])} teams)")

    # Keeper table can also arrive as an overlay while the workbook tab is still empty.
    if not keepers:
        overlay_year = keeper_year
        overlay = load_overlay(f"eligible-keepers-{overlay_year}.json") if overlay_year else None
        if overlay:
            keepers = overlay["players"]
            print(f"  keeper overlay: {overlay_year} ({len(keepers)} players)")

    # ESPN rankings for the upcoming draft, cached by scripts/fetch_rankings.py.
    rankings = None
    if keeper_year:
        rankings = load_overlay(f"espn-rankings-{keeper_year}.json")
    draft_prep = build_draft_prep(rankings, keepers, keeper_year)
    if draft_prep:
        print(
            f"  draft prep: {len(draft_prep['players'])} ranked players, "
            f"{draft_prep['matched_to_rosters']} already rostered"
        )

    player_index = build_player_index(list(boards.values()), roster_years, keepers)
    print(f"  player index: {len(player_index)} players")

    # ---- derived season summary -------------------------------------------------------
    seasons = []
    all_years = sorted(set(standings) | set(stats["seasons"]) | set(roster_years) | set(boards))
    for year in all_years:
        s = standings.get(year, {})
        final_places = {e["place"]: e["owner"] for e in s.get("playoff", [])}
        regular_places = {e["place"]: e["owner"] for e in s.get("regular", [])}
        seasons.append(
            {
                "year": year,
                "champion": final_places.get(1),
                "runner_up": final_places.get(2),
                "third": final_places.get(3),
                "moules": final_places.get(10),
                "regular_season_champion": regular_places.get(1),
                "has_draft": year in boards,
                "has_roster": year in roster_years,
                "has_stats": year in stats["seasons"],
            }
        )

    completed_years = sorted(stats["seasons"].keys())
    latest_completed = completed_years[-1] if completed_years else None
    upcoming = max(boards) if boards else None
    if latest_completed and upcoming and upcoming <= latest_completed:
        upcoming = latest_completed + 1

    latest_final = {}
    if latest_completed and latest_completed in standings:
        latest_final = {e["place"]: e["owner"] for e in standings[latest_completed]["playoff"]}

    bylaws = parse_bylaws(
        bylaws_path,
        os.path.join(args.assets, "rulebook"),
        "assets/rulebook",
        shorten=reg.build_redactor(),
        keep_emails=args.keep_emails,
    )

    teams = []
    for team, info in sorted(reg.teams.items()):
        career = next((c for c in stats["career"] if c["team"] == team), {})
        title = next((t for t in universal.get("titles", []) if t["team"] == team), {})
        teams.append(
            {
                **info,
                "wins": career.get("wins"),
                "losses": career.get("losses"),
                "win_pct": career.get("win_pct"),
                "points_for": career.get("points_for"),
                "championships": title.get("championships"),
                "playoffs": title.get("playoffs"),
            }
        )

    meta = {
        "league": "Original 8 Fantasy Football League",
        "short_name": "O8FFL",
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "sources": {
            "workbook": os.path.basename(workbook_path),
            "bylaws": os.path.basename(bylaws_path) if bylaws_path else None,
        },
        "years": {
            "all": all_years,
            "drafts": sorted(boards),
            "rosters": sorted(roster_years),
            "stats": completed_years,
            "standings": sorted(standings),
            "waivers": waivers["years"],
            "latest_completed": latest_completed,
            "upcoming": upcoming,
            "keepers": keeper_year,
        },
        "teams": teams,
        "seasons": seasons,
    }

    # ---- meeting prep -----------------------------------------------------------------
    keeper_by_team = defaultdict(list)
    for row in keepers:
        if row["team"]:
            keeper_by_team[row["team"]].append(row)

    meeting = {
        "season": upcoming,
        "last_completed": latest_completed,
        "draft_order_selection": draft_order_from_standings(latest_final),
        "final_standings": [
            {"place": p, "owner": o} for p, o in sorted(latest_final.items()) if o
        ],
        "waiver_budgets": [
            {
                "team": t["team"],
                "owner": t["owner"],
                "budget": t["budgets"].get(str(upcoming)) if upcoming else None,
                "left_over": (waivers.get("remaining", {}).get(str(latest_completed), {}) or {}).get(
                    str(t["team"])
                ),
            }
            for t in waivers["teams"]
        ],
        "keeper_counts": [
            {
                "team": team,
                "owner": reg.owner_for_team(team),
                "eligible": sum(1 for r in rows if r["eligible"]),
                "expiring": sum(1 for r in rows if (r["contract_years_remaining"] or 0) == 0),
                "cheapest": sorted(
                    [r for r in rows if r["cost_round"]],
                    key=lambda r: r["cost_round"],
                    reverse=True,
                )[:3],
            }
            for team, rows in sorted(keeper_by_team.items())
        ],
    }

    # ---- write ------------------------------------------------------------------------
    outputs = {
        "meta.json": meta,
        "standings.json": {
            "seasons": {str(y): v for y, v in standings.items()},
            "universal": universal,
        },
        "stats.json": {
            "career": stats["career"],
            "seasons": {str(y): v for y, v in stats["seasons"].items()},
        },
        "keepers.json": {"year": keeper_year, "players": keepers},
        "waivers.json": waivers,
        "players.json": {"players": player_index},
        "meeting.json": meeting,
    }
    if draft_prep:
        outputs["draftprep.json"] = draft_prep
    if bylaws:
        outputs["rulebook.json"] = bylaws

    # Draft boards and rosters are browsed one season at a time, so they get their own
    # files instead of one giant payload. (This is also what a live draft-day board will
    # want: a small, single-season document it can poll or replace.)
    for folder, source in (("drafts", boards), ("rosters", roster_years)):
        os.makedirs(os.path.join(args.out, folder), exist_ok=True)
        for year, payload in source.items():
            outputs[f"{folder}/{year}.json"] = payload

    print("\nwrote:")
    total = 0
    for name, payload in outputs.items():
        size = write_json(os.path.join(args.out, name), payload)
        total += size
        if "/" not in name:
            print(f"  data/{name:<16} {size/1024:7.1f} KB")
    print(f"  data/drafts/*.json + data/rosters/*.json ({len(boards)} + {len(roster_years)} seasons)")
    print(f"  {total/1024:.0f} KB total")

    if reg.unresolved:
        print("\nUnrecognised owner spellings (add them to OWNER_ALIASES in this script):")
        for name, count in reg.unresolved.most_common():
            print(f"  {name!r} x{count}")
    print("\nDone.")


if __name__ == "__main__":
    main()
