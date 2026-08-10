#!/usr/bin/env python3
"""
Build an "<year> Eligible Keepers" table.

Cross-references three things:
  * the previous season's draft board (workbook tab "<year-1> Draft")
  * that season's final rosters (source/final-rosters-<year-1>.json, transcribed from
    the ESPN roster screenshots)
  * the previous keeper tab, for contracts that started before last season

...and applies the league's keeper rules (by-laws 6.1), which are the same rules the
commissioner's own spreadsheet formulas encode:

  Contract years remaining = year_signed + (5 if rookie else 3) - upcoming_year
      A blank "Year Signed" means the contract started last season.
      0 or less  ->  the player returns to the draft pool.

  Keeper round cost   = the round they went in last year's draft, minus one (floored at 1)
      Free-agency pickups cost a 6th. A player acquired by trade keeps the contract and
      the cost of wherever he was originally drafted. A player who was never drafted
      falls back to a 6th.

Anyone not on the final roster is simply absent — you cannot keep a player you dropped.

Usage:  python3 scripts/build_keepers.py [--year 2026]
Writes: source/eligible-keepers-<year>.json   (read by build_data.py)
        build/<year>-eligible-keepers.csv     (paste straight into the workbook tab)
"""

from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from build_data import (  # noqa: E402
    OwnerRegistry,
    as_int,
    as_text,
    cell,
    find_keeper_sheet,
    player_key,
    split_player,
)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Players who entered the NFL in the season just finished. The rookie exception stretches
# a contract from 3 years to 5, so this only changes the countdown, never this year's
# cost. Update it each year (or leave a name out and fix the flag in Excel).
ROOKIE_CLASS = {
    2025: {
        "Ashton Jeanty", "TreVeyon Henderson", "Omarion Hampton", "RJ Harvey",
        "Cam Skattebo", "Kaleb Johnson", "Quinshon Judkins", "Emeka Egbuka",
        "Tetairoa McMillan", "Matthew Golden", "Travis Hunter", "Luther Burden III",
        "Jayden Higgins", "Tyler Warren", "Colston Loveland", "Harold Fannin Jr.",
        "Jaxson Dart", "Dylan Sampson", "Woody Marks", "Devin Neal", "Kyle Monangai",
        "Tahj Brooks", "Jacory Croskey-Merritt", "Tyler Loop", "Oronde Gadsden",
        "Konata Mumpfield", "Tez Johnson", "Bam Knight",
    }
}


# Nicknames the draft board uses that no string-similarity check will ever bridge.
NAME_ALIASES = {
    "jacorycroskeymerritt": "billcroskeymeritt",
}


def load_json(path):
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def solve_keepers(candidates, limit=4):
    """
    Work out which of a team's contract-eligible players were actually kept last year.

    The workbook doesn't record it — the draft board's "Keeper" row is left blank — but it
    is recoverable. A keeper has to be taken at his keeper-cost round, and when two of them
    want the same round the second is pushed to the next free one. So: try every subset of
    candidates (up to the four-keeper limit), simulate that cascade, and keep the largest
    subset whose simulated rounds match what actually happened on the board.

    candidates: [{"key", "cost", "round", ...}]  ->  (kept subset, confident?)
    """
    from itertools import combinations

    best, best_score = [], -1
    for size in range(min(limit, len(candidates)), 0, -1):
        for subset in combinations(candidates, size):
            taken, ok = set(), True
            for player in sorted(subset, key=lambda p: (p["cost"], p["round"])):
                slot = player["cost"]
                while slot in taken:
                    slot += 1
                taken.add(slot)
                if slot != player["round"]:
                    ok = False
                    break
            if not ok:
                continue
            exact = sum(1 for p in subset if p["cost"] == p["round"])
            score = size * 10 + exact
            if score > best_score:
                best, best_score = list(subset), score
        if best:
            break
    return best


def previous_keeper_rows(wb, reg):
    """
    Read last year's keeper tab so contracts that started before last season carry over.
    Returns {player_key: {"year_signed": int|None, "rookie": bool, "owner": str, "cost": int|None}}
    """
    import re

    best = None
    for name in wb.sheetnames:
        m = re.fullmatch(r"(\d{4})\s*(?:eligible\s*)?keepers", name.strip(), re.I)
        if m:
            year = int(m.group(1))
            if best is None or year < best[0]:
                pass
            rows = []
            ws = wb[name]
            header = None
            for r in range(1, 12):
                if (as_text(cell(ws, r, 2)) or "").lower() == "player":
                    header = r
                    break
            if header is None:
                continue
            r = header + 1
            blanks = 0
            while r <= ws.max_row and blanks < 5:
                raw = as_text(cell(ws, r, 2))
                if not raw:
                    blanks += 1
                    r += 1
                    continue
                blanks = 0
                name_only, _, _ = split_player(raw)
                rows.append(
                    {
                        "key": player_key(raw),
                        "player": name_only,
                        "team": as_int(cell(ws, r, 3)),
                        "year_signed": as_int(cell(ws, r, 5)),
                        "rookie": (as_text(cell(ws, r, 6)) or "").lower().startswith("y"),
                        "cost": as_int(cell(ws, r, 9)),
                        "cost_label": as_text(cell(ws, r, 9)),
                    }
                )
                r += 1
            if rows:
                candidate = (year, rows)
                if best is None or candidate[0] > best[0]:
                    best = candidate
    return best or (None, [])


def main():
    ap = argparse.ArgumentParser(description="Build the eligible-keeper table for a season.")
    ap.add_argument("--year", type=int, help="upcoming season (default: the empty keeper tab's year)")
    ap.add_argument("--workbook", help="path to the master workbook")
    args = ap.parse_args()

    import glob

    import openpyxl

    workbook = args.workbook or sorted(
        p for p in glob.glob(os.path.join(ROOT, "*.xlsx")) + glob.glob(os.path.join(ROOT, "source", "*.xlsx"))
        if not os.path.basename(p).startswith("~$")
    )[-1]
    wb = openpyxl.load_workbook(workbook, data_only=True)
    reg = OwnerRegistry()

    tab_year, _ = find_keeper_sheet(wb)
    year = args.year or tab_year
    if not year:
        sys.exit("Could not work out which season to build. Pass --year.")
    prior = year - 1

    print(f"building {year} eligible keepers from the {prior} draft + {prior} final rosters")

    rosters = load_json(os.path.join(ROOT, "source", f"final-rosters-{prior}.json"))
    board = load_json(os.path.join(ROOT, "data", "drafts", f"{prior}.json"))

    # ---- last year's draft: player -> (round, owner, team) --------------------------
    drafted = {}
    for rnd in board["rounds"]:
        for pick in rnd["picks"]:
            if pick.get("player_key"):
                drafted.setdefault(
                    pick["player_key"],
                    {"round": rnd["round"], "team": pick["team"], "owner": pick["owner"], "as_typed": pick["player"]},
                )

    def find_pick(key, team):
        """
        Draft boards are typed by hand and full of near-misses ("Christian McCaffrrey",
        "Tyler Algierr", "Kaimi Fairbaim"). Try the exact key, then the closest spelling.
        """
        hit = drafted.get(key) or drafted.get(NAME_ALIASES.get(key, ""))
        if hit:
            return hit, ("board calls him “%s”" % hit["as_typed"]) if key in NAME_ALIASES else None
        best, best_score = None, 0.0
        for candidate, pick in drafted.items():
            score = difflib.SequenceMatcher(None, key, candidate).ratio()
            if pick["team"] == team:
                score += 0.06  # same roster is a tiebreaker, not a licence
            if score > best_score:
                best, best_score = pick, score
        if best and best_score >= 0.84:
            return best, f"matched to “{best['as_typed']}” on the board (round {best['round']})"
        return None, None

    # Picks made with a traded selection live in the board's note rows rather than a round
    # row: one line declaring the pick ("Round 9 via Ryan"), then the player taken with it.
    # Read them per team, in the order they were typed, and pair them up.
    acquired_picks, note_players = {}, {}
    for note in board.get("notes", []):
        for entry in note["entries"]:
            team = entry["team"]
            text = (entry["text"] or "").strip()
            m = re.match(r"round\s*(\d+)", text, re.I)
            if m:
                acquired_picks.setdefault(team, []).append(int(m.group(1)))
            elif text:
                note_players.setdefault(team, []).append(text)

    traded_pick_rounds = {}
    for team, players in note_players.items():
        rounds = acquired_picks.get(team, [])
        for i, who in enumerate(players):
            if i < len(rounds):
                traded_pick_rounds[(team, player_key(who))] = rounds[i]

    # ---- previous keeper tab: contracts already running -----------------------------
    prev_year, prev_rows = previous_keeper_rows(wb, reg)
    prev_by_key = {row["key"]: row for row in prev_rows}

    def find_prev(key):
        match = difflib.get_close_matches(key, prev_by_key.keys(), n=1, cutoff=0.86)
        return prev_by_key[match[0]] if match else None
    print(f"  carried {len(prev_by_key)} rows forward from the {prev_year} keeper tab")

    # ---- who did each team actually keep last year? ---------------------------------
    from collections import defaultdict

    # Excel has already re-evaluated last year's tab against this year, so some rows now
    # read "Return to Pool" even though the contract was alive last season. Rebuild those
    # costs from the draft before last, the same way the sheet's own formula does.
    older_board = {}
    older_path = os.path.join(ROOT, "data", "drafts", f"{prior - 1}.json")
    if os.path.exists(older_path):
        for rnd in load_json(older_path)["rounds"]:
            for pick in rnd["picks"]:
                if pick.get("player_key"):
                    older_board.setdefault(pick["player_key"], rnd["round"])

    candidates = defaultdict(list)
    for key, row in prev_by_key.items():
        signed = row["year_signed"] or (prior - 1)
        if signed + (5 if row["rookie"] else 3) - prior <= 0:
            continue  # contract was already up; any pick last year started a new one
        pick, _ = find_pick(key, row["team"])
        if not pick:
            continue
        if not row.get("cost"):
            older = older_board.get(key)
            row["cost"] = max(1, older - 1) if older else 6
        if not row.get("cost"):
            continue
        candidates[pick["team"]].append(
            {"key": key, "player": row["player"], "cost": row["cost"], "round": pick["round"], "team": pick["team"]}
        )

    kept_keys = {}
    for team, rows in candidates.items():
        for player in solve_keepers(rows):
            kept_keys[player["key"]] = player
    print(f"  reconstructed {len(kept_keys)} keepers from the {prior} draft board")

    rookies = ROOKIE_CLASS.get(prior, set())
    out_rows = []
    review = []

    for block in rosters["teams"]:
        team = block["team"]
        owner = reg.owner_for_team(team)
        for slot, raw_name, nfl_team, position, acq in block["players"]:
            if not raw_name or raw_name.lower() == "empty":
                continue
            key = player_key(raw_name)
            name, _, _ = split_player(raw_name)
            pick, fuzzy = find_pick(key, team)
            prev = prev_by_key.get(key)
            if prev is None:
                prev = find_prev(key)

            # --- when did this contract start? -------------------------------------
            # A player who was on last year's keeper list AND was taken in last year's
            # draft by the same owner was kept, so his contract predates last season.
            year_signed = None
            rookie = name in rookies
            kept_last_year = key in kept_keys
            if kept_last_year:
                # Blank on last year's tab meant "signed the year before that".
                year_signed = prev["year_signed"] or (prior - 1)
                rookie = prev["rookie"]

            effective_signed = year_signed or prior
            remaining = effective_signed + (5 if rookie else 3) - year

            # --- what does he cost? -------------------------------------------------
            if remaining <= 0:
                cost_label = "Return to Pool"
            elif acq == "Free Agency":
                cost_label = "6"
            elif pick:
                cost_label = str(max(1, pick["round"] - 1))
            else:
                # Drafted or traded for, but absent from every round row — usually a pick
                # made with a traded selection, which the board logs as a note instead.
                traded_round = traded_pick_rounds.get((team, key))
                if traded_round:
                    cost_label = str(max(1, traded_round - 1))
                    review.append(
                        {
                            "owner": owner,
                            "player": name,
                            "acq": acq,
                            "reason": f"taken with a pick traded in (board note: round {traded_round})",
                            "assumed": cost_label,
                        }
                    )
                else:
                    cost_label = "6"
                    review.append(
                        {
                            "owner": owner,
                            "player": name,
                            "acq": acq,
                            "reason": f"no match in the {prior} draft board",
                            "assumed": cost_label,
                        }
                    )

            if fuzzy:
                review.append({"owner": owner, "player": name, "acq": acq, "reason": fuzzy, "assumed": cost_label})

            out_rows.append(
                {
                    "player": name,
                    "player_key": key,
                    "nfl_team": nfl_team,
                    "position": position,
                    "team": team,
                    "owner": owner,
                    "year_signed": year_signed,
                    "rookie": rookie,
                    "contract_years_remaining": max(0, remaining),
                    "acquired": acq,
                    "cost_round": int(cost_label) if cost_label.isdigit() else None,
                    "cost_label": cost_label,
                    "eligible": cost_label != "Return to Pool",
                    "kept_last_year": kept_last_year,
                    "prior_round": pick["round"] if pick else None,
                }
            )

    out_rows.sort(key=lambda r: (r["team"], r["player"]))

    # ---- write -----------------------------------------------------------------------
    out_path = os.path.join(ROOT, "source", f"eligible-keepers-{year}.json")
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(
            {
                "year": year,
                # Who was kept in the season just gone, so the draft board can flag them.
                "prior_keepers": [
                    {"team": p["team"], "player": p["player"], "player_key": p["key"], "round": p["round"]}
                    for p in sorted(kept_keys.values(), key=lambda p: (p["team"], p["round"]))
                ],
                "generated_from": {
                    "draft": f"{prior} Draft tab",
                    "rosters": f"source/final-rosters-{prior}.json",
                    "previous_keepers": f"{prev_year} Eligible Keepers tab",
                },
                "players": out_rows,
            },
            fh,
            ensure_ascii=False,
            indent=1,
        )

    # The CSV is for pasting into Excel, so it carries the workbook's full names.
    full_names = {display: full for full, display in reg.full_to_display.items()}

    os.makedirs(os.path.join(ROOT, "build"), exist_ok=True)
    csv_path = os.path.join(ROOT, "build", f"{year}-eligible-keepers.csv")
    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            ["Player", "Team #", "Current Owner", "Year Signed", "Rookie?",
             "Contract Years Remaining", "Acq?", "Keeper Round Cost"]
        )
        for row in out_rows:
            writer.writerow(
                [
                    row["player"],
                    row["team"],
                    full_names.get(row["owner"], row["owner"]),
                    row["year_signed"] or "",
                    "Yes" if row["rookie"] else "",
                    row["contract_years_remaining"],
                    row["acquired"],
                    row["cost_label"],
                ]
            )

    review_path = os.path.join(ROOT, "build", f"{year}-keeper-review.md")
    with open(review_path, "w", encoding="utf-8") as fh:
        fh.write(f"# {year} keeper table — what to check\n\n")
        fh.write(
            f"Generated from the {prior} draft board, the {prior} final rosters and the "
            f"{prev_year} keeper tab.\n\n"
            "**Keeper round cost is rule-derived and should be right.** It is last year's "
            "draft round minus one, a 6th for a free-agent pickup, and a traded player "
            "keeps his original cost.\n\n"
            "**Contract years remaining involves one guess.** The workbook never records "
            "who was actually kept — the draft board's Keeper row is blank — so it is "
            "reconstructed by finding the players taken at exactly their keeper cost "
            "(allowing for stacking). If one of these is wrong, fix the Year Signed cell "
            "in Excel and everything downstream follows.\n\n"
        )
        fh.write(f"## Reconstructed {prior} keepers ({len(kept_keys)})\n\n")
        for team in sorted({p["team"] for p in kept_keys.values()}):
            names = [p for p in kept_keys.values() if p["team"] == team]
            fh.write(f"- **{reg.owner_for_team(team)}** — " + ", ".join(
                f"{p['player']} (cost R{p['cost']}, taken R{p['round']})" for p in sorted(names, key=lambda p: p["cost"])
            ) + "\n")
        if review:
            fh.write(f"\n## Players needing a human eye ({len(review)})\n\n")
            for item in review:
                fh.write(f"- **{item['owner']} — {item['player']}** ({item['acq']}): {item['reason']} → using round {item['assumed']}\n")
        fh.write(f"\n## Rookie flags applied\n\n")
        rookie_rows = [r for r in out_rows if r["rookie"] and not r["year_signed"]]
        fh.write(", ".join(sorted(r["player"] for r in rookie_rows)) + "\n\n")
        fh.write("Rookie contracts run five years instead of three, so a wrong flag changes "
                 "the countdown but never this year's cost.\n")

    keepable = [r for r in out_rows if r["eligible"]]
    print(f"  {len(out_rows)} rostered players, {len(keepable)} keepable, "
          f"{len(out_rows) - len(keepable)} returning to the pool")
    print(f"  wrote {os.path.relpath(out_path, ROOT)}")
    print(f"  wrote {os.path.relpath(csv_path, ROOT)}")
    print(f"  wrote {os.path.relpath(review_path, ROOT)}")

    if review:
        print(f"\n  {len(review)} player(s) need a human eye:")
        for item in review:
            print(f"    {item['owner']:<8} {item['player']:<26} {item['acq']:<12} {item['reason']} → assumed round {item['assumed']}")

    print("\n  per team:")
    for team in sorted({r["team"] for r in out_rows}):
        rows = [r for r in out_rows if r["team"] == team]
        cheap = sorted([r for r in rows if r["cost_round"]], key=lambda r: r["cost_round"])[:4]
        pool = [r["player"] for r in rows if not r["eligible"]]
        print(f"    {reg.owner_for_team(team):<8} {len(rows):>2} players, "
              f"{len(rows) - len(pool):>2} keepable | best: " + ", ".join(f"{r['player']} (R{r['cost_round']})" for r in cheap))
        if pool:
            print(f"             back to the pool: {', '.join(pool)}")


if __name__ == "__main__":
    main()
